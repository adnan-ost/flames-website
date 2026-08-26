"""
Fill the Blink POS import workbooks from the live menu.

    python3 scripts/export-pos.py            # both sheets, default paths

Reads Sanity, not src/data/menu.ts, so the POS and the website cannot drift:
whatever the restaurant last edited in the Studio is what lands here. Re-run
it after a menu change and the sheet is rebuilt.

Blink's model is Category -> Item -> Variation (see the architecture PDF), so
a dish sold in two sizes becomes two rows sharing an item_name, differing by
`label` and `item_price`.

Deliberately left empty, because we do not have the facts:
  buying_price    — cost per dish is the restaurant's, and nobody has given it
  discount_price  — no discounts are defined anywhere in this project
Guessing either would put invented numbers into the system that prices real
orders and computes real margins.
"""
import json, sys, urllib.parse, urllib.request
from pathlib import Path

import openpyxl

PROJECT, DATASET, API = "byr90f6b", "production", "v2026-08-24"

QUERY = """*[_type == "menuSection" && defined(slug.current)] | order(order asc, title asc){
  "section": title, "filter": filter, "intro": intro,
  "items": items[]->{
    name, description, price,
    "sizes": sizes[]{label, price},
    "image": image.asset->url
  }
}"""

# Single-size items still need a variation row; Blink expects one per item.
DEFAULT_LABEL = "Regular"
STATUS = "active"


def fetch():
    url = (f"https://{PROJECT}.api.sanity.io/{API}/data/query/{DATASET}"
           f"?query={urllib.parse.quote(QUERY)}")
    with urllib.request.urlopen(url, timeout=60) as r:
        return json.load(r)["result"]


def rows(sections):
    out, categories = [], []
    for index, section in enumerate(sections, start=1):
        categories.append([index, section["section"], section.get("filter") or "",
                           (section.get("intro") or "").strip()])
        for item in section.get("items") or []:
            if not item or not item.get("name"):
                continue
            sizes = [s for s in (item.get("sizes") or [])
                     if s and s.get("label") and s.get("price")]
            variations = ([(s["label"], s["price"]) for s in sizes] if len(sizes) > 1
                          else [(DEFAULT_LABEL, item.get("price"))])
            for label, price in variations:
                out.append([
                    index,                      # category_id
                    item["name"],               # item_name
                    label,                      # label (variation)
                    price if price else None,   # item_price, pre-tax
                    None,                       # buying_price  — not ours to invent
                    None,                       # discount_price — none defined
                    STATUS,                     # status
                    item.get("image") or "",    # image_link
                    (item.get("description") or "").strip(),
                ])
    return out, categories


def reset_sheet(ws):
    """
    Clear a template sheet down to its header row.

    Unmerging first is not optional. Blink's category template ends with a
    merged "remove the sample data" note spanning A:C; deleting the rows
    leaves the merged range behind, and openpyxl then silently drops any write
    to the covered cells — which is how a category quietly lost its priority
    and description while the row above and below looked perfect.
    """
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def write_categories(target: Path, categories):
    """
    Fill Blink's category template: Category Name | Priority | Description.

    Priority is the menu's own order, and it is the same 1..20 used as
    category_id in the item sheet — so whichever order Blink assigns its ids
    in, the two sheets describe the same sequence.

    The template ships with a sample row and a note telling you to delete it;
    both are cleared here rather than left for someone to notice.
    """
    wb = openpyxl.load_workbook(target)
    ws = wb["Worksheet"] if "Worksheet" in wb.sheetnames else wb.worksheets[0]
    reset_sheet(ws)
    for cid, name, _filter, description in categories:
        ws.append([name, cid, description])
    wb.save(target)
    print(f"  {len(categories)} categories written to {target}")


def main():
    target = Path(sys.argv[1] if len(sys.argv) > 1 else "Blink POS Data/Blink POS Data.xlsx")
    category_target = Path("Blink POS Data/Category_Bulk_Upload_Template.xlsx")
    sections = fetch()
    data, categories = rows(sections)

    wb = openpyxl.load_workbook(target)
    ws = wb["Worksheet"] if "Worksheet" in wb.sheetnames else wb.worksheets[0]

    # Rebuild the rows under the existing header, leaving the header untouched.
    reset_sheet(ws)
    for row in data:
        ws.append(row)

    # The ids above are ours, in menu order. Blink's own category ids go here
    # if they differ, so the mapping is never guessed at in the item sheet.
    name = "Categories (reference)"
    if name in wb.sheetnames:
        del wb[name]
    ref = wb.create_sheet(name)
    ref.append(["category_id", "category_name", "website_filter", "description"])
    for c in categories:
        ref.append(c)

    wb.save(target)
    if category_target.exists():
        write_categories(category_target, categories)

    sized = sum(1 for r in data if r[2] != DEFAULT_LABEL)
    print(f"  {len(data)} rows across {len(categories)} categories")
    print(f"  {sized} are size variations; {len(data) - sized} single-size")
    print(f"  written to {target}")


if __name__ == "__main__":
    main()
